from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill
import time
from datetime import datetime


def convert_date_format(date_str):
    # Convertendo a string da data para um objeto datetime
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')

    # Formatando o objeto datetime para o formato desejado
    return date_obj.strftime('%d/%m')


def minera_rec(salvar=False):
    CHROME_DRIVER_PATH = '/usr/bin/chromedriver'

    options = Options()
    options.binary_location = '/home/diogo/Downloads/chrome-linux64/chrome'

    options.executable_path = CHROME_DRIVER_PATH

    browser = webdriver.Chrome(options=options)

    url = "https://recnplay.pe/programacao-2023"
    browser.get(url)

    # Dá tempo para a página carregar - ajuste conforme necessário
    browser.implicitly_wait(30)
    print("\033[93m", 'Aguardando carregar dados 5, 4, 3, 2, 1...')
    time.sleep(5)

    # Localizar todos os elementos de opção de data
    date_options = browser.find_elements(By.CSS_SELECTOR, "input.btn-check.date-radio")
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Data', 'Horário', 'Nome do Evento', 'Formato do Evento', 'Local do Evento', 'Speakers']
    ws.append(headers)  # Headers da planilha

    # Formatar cabeçalho
    for cell in ws["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")  # Texto em negrito e letra branca
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Fundo preto

    colors = {
        18: "90EE90",  # verde claro
        19: "FFFFE0",  # amarelo claro
        20: "ADD8E6",  # azul claro
        21: "FFB6C1"  # vermelho claro
    }

    for dia, option in enumerate(date_options, start=1):
        print("\033[94m", f"Coletando dados do {dia}º dia... ")
        date_str = option.get_attribute("value")
        date = convert_date_format(date_str)

        # Coleta os dados do primeiro dia sem clicar, e clica nos dias seguintes
        if dia > 1:

            if option.get_attribute('id') == 'option4':
                label_selector = f"labe[for='{option.get_attribute('id')}']"
            else:
                label_selector = f"label[for='{option.get_attribute('id')}']"

            # Buscar o elemento <label> associado e clicar nele
            label_for_option = browser.find_element(By.CSS_SELECTOR, label_selector)
            label_for_option.click()
            time.sleep(5)  # Aguarde a página ser atualizada

        soup = BeautifulSoup(browser.page_source, 'html.parser')
        eventos = soup.find('div', {'id': 'result-container'}).find_all('div', {'class': 'card-programacao'})

        for evento in eventos:
            horario = evento.find('p').text.strip()
            nome_evento = evento.find('p', {'class': 'title-programacao'}).text.strip()
            formato_evento = evento.find('p', {'id': 'formato'}).text.strip()
            local_evento = evento.find('p', {'id': 'local'}).text.strip()
            speakers = evento.find('p', {'id': 'speakers'}).text.strip()

            ws.append([date, horario, nome_evento, formato_evento, local_evento, speakers])

        if dia == 4:
            break

    # Colorir as linhas com base na data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        day = int(row[0].value.split("/")[0])  # row[0] se refere à primeira célula da linha (coluna de datas)
        color_key = day if day in colors else 18  # Default para 1 (verde claro) caso não encontre a cor
        fill = PatternFill(start_color=colors[color_key], end_color=colors[color_key], fill_type="solid")
        for cell in row:
            cell.fill = fill

    # Configurar manualmente a largura de cada coluna
    column_widths = [10, 15, 60, 20, 30, 50]  # Ajuste esses valores conforme necessário
    for i, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = width

    # Salva os dados no Excel
    if salvar:
        print("\033[93m", 'Salvando dados na planilha...')
        wb.save('programacao_recnplay.xlsx')
        print("\033[92m", 'Dados Salvos!')

    print("\033[91m", 'Fechando o browser. Good Bye!')
    browser.quit()


if __name__ == '__main__':
    minera_rec(salvar=True)
